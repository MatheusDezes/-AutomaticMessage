import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import customtkinter as ctk
import tkinter as tk
from threading import Thread

# Vari치vel de controle para interromper a automa칞칚o
automation_running = False

# Fun칞칚o de automa칞칚o
def start_automation(text_widget):
    global automation_running
    automation_running = True  # Ativar a automa칞칚o

    try:
        # Abrir o WhatsApp Web
        log_message(text_widget, "Abrindo WhatsApp Web...")
        webbrowser.open('https://web.whatsapp.com/')
        sleep(30)  # Tempo para escanear o QR Code

        # Ler planilha
        workbook = openpyxl.load_workbook('clientes.xlsx')
        pageClients = workbook['Planilha1']

        for idx, line in enumerate(pageClients.iter_rows(min_row=2), start=2):
            # Verificar se a automa칞칚o foi interrompida
            if not automation_running:
                log_message(text_widget, "Automa칞칚o interrompida pelo usu치rio.")
                break

            # Armazenar dados: nome, telefone, CashBack
            name = line[0].value
            phone = line[1].value
            cashBack = line[2].value
            vencimento = pageClients['F2'].value

            # Verifica칞칚o se todos os dados (nome, telefone, cashback) est칚o vazios
            if not name and not phone and not cashBack:
                log_message(text_widget, "Fim do envio. Todas as c칠lulas est칚o vazias.")
                pageClients.cell(row=idx, column=4, value='Fim da execu칞칚o')
                workbook.save('clientes.xlsx')
                break

            # Verifica칞칚o b치sica se dados est칚o presentes
            if not name or not phone or not cashBack:
                log_message(text_widget, f"Dados faltando para {name}. Pulando...")
                pageClients.cell(row=idx, column=4, value='Dados incompletos')
                continue

            # Mensagem a ser enviada
            message = f'Ol치 {name}, tudo bem? Polo Wear SP Market, passando para lembra-la que voc칡 tem um valor de desconto em seu cashback de R${cashBack} vinculado ao seu CPF,  venha resgatar, voc칡 s칩 precisa fazer uma compra do dobro do valor do b칪nus 游땎, o desconto ser치 abatido no m치ximo de 50%, da sua compra, o mesmo expira em {vencimento.strftime("%d/%m/%y")}.' 

            try:
                # Gerar link do WhatsApp com a mensagem
                linkMessageWhatsapp = f'https://web.whatsapp.com/send?phone={phone}&text={quote(message)}'
                webbrowser.open(linkMessageWhatsapp)
                sleep(7)  # Tempo para carregar a janela do WhatsApp com o n칰mero

                # Verificar se a automa칞칚o foi interrompida
                if not automation_running:
                    log_message(text_widget, "Automa칞칚o interrompida pelo usu치rio.")
                    break

                sleep(2)
                pyautogui.hotkey('Enter')  # Fechar a aba ap칩s o envio
                sleep(5)
                pyautogui.hotkey('ctrl', 'w')  # Fechar a aba ap칩s o envio
                pageClients.cell(row=idx, column=4, value='Sucesso ao enviar mensagem')
                log_message(text_widget, f"Mensagem enviada para {name}")

            except Exception as e:
                log_message(text_widget, f"Erro ao enviar mensagem para {name}")
                pageClients.cell(row=idx, column=4, value='Erro ao enviar mensagem')

        try:
            workbook.save('clientes.xlsx')
            log_message(text_widget, "Arquivo salvo com sucesso.")
        except Exception:
            log_message(text_widget, "Erro ao salvar arquivo.")
    except Exception as e:
        log_message(text_widget, f"Erro durante a execu칞칚o: {e}")

# Fun칞칚o para logar mensagens no widget de texto
def log_message(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)  # Scroll autom치tico para o final

# Fun칞칚o para encerrar a automa칞칚o
def stop_automation():
    global automation_running
    automation_running = False  # Desativar a automa칞칚o

# Fun칞칚o para iniciar a automa칞칚o em um thread separado
def start_thread(text_widget):
    global automation_running
    if not automation_running:  # Evitar que m칰ltiplas inst칙ncias sejam iniciadas
        thread = Thread(target=start_automation, args=(text_widget,))
        thread.start()

# Criando a interface gr치fica com CustomTkinter
app = ctk.CTk()
app.title("PoloBot Automa칞칚o")
app.geometry("600x400")
app.iconbitmap("./fig/iconBot.ico")
app.resizable(False, False)

# 츼rea de log de sa칤da
text_widget = tk.Text(app, height=15, width=70)
text_widget.pack(pady=10)

# Bot칚o para iniciar a automa칞칚o
start_button = ctk.CTkButton(app, text="Iniciar PoloBot", command=lambda: start_thread(text_widget))
start_button.pack(pady=10)

# Bot칚o para encerrar a automa칞칚o
stop_button = ctk.CTkButton(app, text="Encerrar PoloBot", command=stop_automation)
stop_button.pack(pady=10)

# Rodando a aplica칞칚o
app.mainloop()