import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import customtkinter as ctk
import tkinter as tk
from threading import Thread

# Variável de controle para interromper a automação
automation_running = False

# Função de automação
def start_automation(text_widget):
    global automation_running
    automation_running = True  # Ativar a automação

    try:
        # Abrir o WhatsApp Web
        log_message(text_widget, "Abrindo WhatsApp Web...")
        webbrowser.open('https://web.whatsapp.com/')
        sleep(30)  # Tempo para escanear o QR Code

        # Ler planilha
        workbook = openpyxl.load_workbook('clientes.xlsx')
        pageClients = workbook['Planilha1']

        for idx, line in enumerate(pageClients.iter_rows(min_row=2), start=2):
            # Verificar se a automação foi interrompida
            if not automation_running:
                log_message(text_widget, "Automação interrompida pelo usuário.")
                break

            # Armazenar dados: nome, telefone, CashBack
            name = line[0].value
            phone = line[1].value
            cashBack = line[2].value
            vencimento = pageClients['F2'].value

            # Verificação se todos os dados (nome, telefone, cashback) estão vazios
            if not name and not phone and not cashBack:
                log_message(text_widget, "Fim do envio. Todas as células estão vazias.")
                pageClients.cell(row=idx, column=4, value='Fim da execução')
                workbook.save('clientes.xlsx')
                break

            # Verificação básica se dados estão presentes
            if not name or not phone or not cashBack:
                log_message(text_widget, f"Dados faltando para {name}. Pulando...")
                pageClients.cell(row=idx, column=4, value='Dados incompletos')
                continue

            # Mensagem a ser enviada
            message = f'Olá {name}, tudo bem? Polo Wear SP Market, passando para lembra-la que você tem um valor de desconto em seu cashback de R${cashBack} vinculado ao seu CPF,  venha resgatar, você só precisa fazer uma compra do dobro do valor do bônus 😃, o desconto será abatido no máximo de 50%, da sua compra, o mesmo expira em {vencimento.strftime("%d/%m/%y")}.' 

            try:
                # Gerar link do WhatsApp com a mensagem
                linkMessageWhatsapp = f'https://web.whatsapp.com/send?phone={phone}&text={quote(message)}'
                webbrowser.open(linkMessageWhatsapp)
                sleep(10)  # Tempo para carregar a janela do WhatsApp com o número

                # Verificar se a automação foi interrompida
                if not automation_running:
                    log_message(text_widget, "Automação interrompida pelo usuário.")
                    break

                # Tentar localizar a seta de envio da mensagem
                arrow = pyautogui.locateCenterOnScreen('arrow.png')

                if arrow:
                    sleep(5)
                    pyautogui.click(arrow[0], arrow[1])  # Clicar na seta para enviar
                    sleep(5)
                    pyautogui.hotkey('ctrl', 'w')  # Fechar a aba após o envio
                    pageClients.cell(row=idx, column=4, value='Sucesso ao enviar mensagem')
                    log_message(text_widget, f"Mensagem enviada para {name}")
                else:
                    log_message(text_widget, f"Seta de envio não encontrada para {name}.")
                    pageClients.cell(row=idx, column=4, value='Seta de envio não encontrada')

            except Exception as e:
                log_message(text_widget, f"Erro ao enviar mensagem para {name}")
                pageClients.cell(row=idx, column=4, value='Erro ao enviar mensagem')

        try:
            workbook.save('clientes.xlsx')
            log_message(text_widget, "Arquivo salvo com sucesso.")
        except Exception:
            log_message(text_widget, "Erro ao salvar arquivo.")
    except Exception as e:
        log_message(text_widget, f"Erro durante a execução: {e}")

# Função para logar mensagens no widget de texto
def log_message(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)  # Scroll automático para o final

# Função para encerrar a automação
def stop_automation():
    global automation_running
    automation_running = False  # Desativar a automação

# Função para iniciar a automação em um thread separado
def start_thread(text_widget):
    global automation_running
    if not automation_running:  # Evitar que múltiplas instâncias sejam iniciadas
        thread = Thread(target=start_automation, args=(text_widget,))
        thread.start()

# Criando a interface gráfica com CustomTkinter
app = ctk.CTk()
app.title("PoloBot WhatsApp")
app.geometry("600x400")

# Área de log de saída
text_widget = tk.Text(app, height=15, width=70)
text_widget.pack(pady=10)

# Botão para iniciar a automação
start_button = ctk.CTkButton(app, text="Iniciar PoloBot", command=lambda: start_thread(text_widget))
start_button.pack(pady=10)

# Botão para encerrar a automação
stop_button = ctk.CTkButton(app, text="Encerrar PoloBot", command=stop_automation)
stop_button.pack(pady=10)

# Rodando a aplicação
app.mainloop()

